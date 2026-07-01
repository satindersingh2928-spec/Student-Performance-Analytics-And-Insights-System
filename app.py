# ==============================
# 🔹 Standard Library Imports
# ==============================
import os
import re
from io import BytesIO
from datetime import datetime
import secrets

# ==============================
# 🔹 Third-Party Imports
# ==============================
from dotenv import load_dotenv

from flask import Flask, render_template, request, redirect, session, flash,send_file, url_for

from werkzeug.utils import secure_filename
from werkzeug.security import generate_password_hash, check_password_hash
import uuid

# ReportLab (PDF)
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph
)
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# OpenPyXL (Excel)
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

# ==============================
# 🔹 Local Imports
# ==============================
from models.db import *

# ==============================
# 🔹 Environment Variables
# ==============================
load_dotenv()
ADMIN_SECRET_KEY = os.getenv("ADMIN_SECRET_KEY")

app = Flask(__name__)

UPLOAD_FOLDER = "uploads"
app.config["UPLOAD_FOLDER"] = UPLOAD_FOLDER

app.secret_key = os.getenv("SECRET_KEY")

from config import DB_CONFIG

def get_connection():
    return mysql.connector.connect(**DB_CONFIG)

@app.route("/")
def home():
    return render_template("home.html")

@app.route("/register", methods=["GET", "POST"])
def register():

    if request.method == "POST":

        # 🔹 Get form data
        full_name = request.form.get("full_name", "").strip()
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "")
        confirm_password = request.form.get("confirm_password", "")
        secret_key = request.form.get("secret_key", "")

        # 🔐 Basic validation
        if not all([full_name, username, password, confirm_password, secret_key]):
            flash("All fields are required.", "danger")
            return redirect("/register")

        if password != confirm_password:
            flash("Passwords do not match.", "danger")
            return redirect("/register")

        if len(password) < 6:
            flash("Password must be at least 6 characters.", "danger")
            return redirect("/register")

        if secret_key != ADMIN_SECRET_KEY:
            flash("Invalid Admin Secret Key.", "danger")
            return redirect("/register")

        try:
            conn = get_connection()
            cursor = conn.cursor(dictionary=True)

            # 🔍 Check existing user
            cursor.execute(
                "SELECT id FROM users WHERE username=%s",
                (username,)
            )
            if cursor.fetchone():
                flash("Username already exists.", "danger")
                return redirect("/register")

            # 🔐 Hash password
            hashed_password = generate_password_hash(password)

            # 📝 Insert user
            cursor.execute("""
                INSERT INTO users (
                    full_name,
                    username,
                    password,
                    role,
                    student_id
                )
                VALUES (%s, %s, %s, 'admin', NULL)
            """, (full_name, username, hashed_password))

            conn.commit()

            # 🧾 Log activity
            log_activity(
                username,
                "Register Admin",
                f"New admin account created: {username}"
            )

            flash("Admin registered successfully. Please login.", "success")
            return redirect("/login")

        except Exception as e:
            return f"Error: {str(e)}"

        finally:
            cursor.close()
            conn.close()

    return render_template("register.html")

@app.route("/login", methods=["GET", "POST"])
def login():

    if request.method == "POST":

        # 🔹 Get form data
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "")

        # 🔐 Basic validation
        if not username or not password:
            flash("Username and Password are required.", "danger")
            return redirect("/login")

        try:
            conn = get_connection()
            cursor = conn.cursor(dictionary=True)

            cursor.execute(
                "SELECT * FROM users WHERE username = %s",
                (username,)
            )

            user = cursor.fetchone()

            if user and check_password_hash(user["password"], password):

                # 🔐 Session set
                session.clear()
                session["username"] = user["username"]
                session["role"] = user["role"]

                # 🧾 Log activity
                log_activity(
                    user["username"],
                    "Login",
                    f"{user['role'].capitalize()} logged into the system"
                )

                # 🎯 Redirect based on role
                if user["role"] == "admin":
                    flash("Welcome Admin!", "success")
                    return redirect("/dashboard")
                else:
                    flash("Login Successful!", "success")
                    return redirect("/student_dashboard")

            else:
                flash("Invalid Username or Password", "danger")
                return redirect("/login")

        except Exception as e:
            return f"Error: {str(e)}"

        finally:
            cursor.close()
            conn.close()

    return render_template("login.html")

@app.route("/dashboard")
def dashboard():
    import time
    start = time.perf_counter()

    # 🔐 Session check
    if "username" not in session:
        return redirect("/login")

    # 🔗 DB Connection
    conn = get_connection()
    cursor = conn.cursor()

    # 📊 Basic Stats
    total_students = get_total_students(cursor)
    avg_percentage = get_average_percentage(cursor)
    avg_attendance = get_average_attendance(cursor)
    pass_count = get_pass_count(cursor)
    fail_count = get_fail_count(cursor)

    # 📌 Attendance Insights
    highest_attendance_student = get_highest_attendance_student(cursor)
    lowest_attendance_student = get_lowest_attendance_student(cursor)

    # 🏆 Performance Insights
    top_performers = get_top_performers(cursor)
    bottom_performers = get_bottom_performers(cursor)

    # 🔚 Close DB
    cursor.close()
    conn.close()

    print("Dashboard time:", round(time.perf_counter() - start, 3), "sec")

    # 🎨 Render Template
    return render_template(
        "dashboard.html",
        username=session["username"],
        total_students=total_students,
        avg_percentage=avg_percentage,
        avg_attendance=avg_attendance,
        pass_count=pass_count,
        fail_count=fail_count,
        highest_attendance_student=highest_attendance_student,
        lowest_attendance_student=lowest_attendance_student,
        top_performers=top_performers,
        bottom_students=bottom_performers
    )
    
@app.route("/logout")
def logout():

    session.clear()

    return redirect("/login")
    
@app.route("/students")
def students():

    # 🔐 Auth check
    if "username" not in session:
        return redirect("/login")

    if session.get("role") != "admin":
        return redirect("/student_dashboard")

    # 🔎 Filters
    search = request.args.get("search", "").strip()
    course = request.args.get("course", "").strip()
    page = request.args.get("page", 1, type=int)
    per_page = 10
    offset = (page - 1) * per_page

    conn = get_connection()
    cursor = conn.cursor(dictionary=True)

    # 🔥 Base Query
    base_query = "FROM students WHERE 1=1"
    params = []

    # 🔎 Dynamic Filters (NO DUPLICATION)
    if search:
        base_query += " AND student_name LIKE %s"
        params.append(f"%{search}%")

    if course:
        base_query += " AND course = %s"
        params.append(course)

    # 📊 Total Count
    count_query = "SELECT COUNT(*) AS total " + base_query
    cursor.execute(count_query, params)
    total_students = cursor.fetchone()["total"]

    # 📄 Data Query with Pagination
    data_query = "SELECT * " + base_query + " ORDER BY student_id LIMIT %s OFFSET %s"
    cursor.execute(data_query, params + [per_page, offset])
    students = cursor.fetchall()

    total_pages = (total_students + per_page - 1) // per_page

    cursor.close()
    conn.close()

    return render_template(
        "students.html",
        students=students,
        search=search,
        course=course,
        page=page,
        total_pages=total_pages
    )

@app.route("/export_pdf")
def export_pdf():

    if "username" not in session:
        return redirect("/login")

    if session.get("role") != "admin":
        return redirect("/student_dashboard")

    try:
        conn = get_connection()
        cursor = conn.cursor(dictionary=True)

        cursor.execute("""
            SELECT
                registration_no, student_name, gender, age, city,
                course, specialization, year, semester,
                attendance_percentage, assignment_marks,
                internal_marks, practical_marks, final_exam_marks,
                total_marks, percentage, grade
            FROM students
            ORDER BY student_id
        """)

        students = cursor.fetchall()
        cursor.close()
        conn.close()

        total_students = len(students)
        current_date = datetime.now().strftime("%d-%m-%Y %I:%M %p")

        filename = f"student_report_{int(datetime.now().timestamp())}.pdf"
        pdf_path = os.path.join("static", filename)

        doc = SimpleDocTemplate(
            pdf_path,
            pagesize=landscape(A4),
            rightMargin=15,
            leftMargin=15,
            topMargin=25,
            bottomMargin=30
        )

        styles = getSampleStyleSheet()
        elements = []

        # 🔥 TITLE
        elements.append(Paragraph("<b>Student Performance Analytics System</b>", styles["Title"]))
        elements.append(Paragraph("<b>Complete Student Report</b>", styles["Heading2"]))
        elements.append(Spacer(1, 10))

        elements.append(Paragraph(f"<b>Total Students :</b> {total_students}", styles["Normal"]))
        elements.append(Paragraph(f"<b>Generated :</b> {current_date}", styles["Normal"]))
        elements.append(Spacer(1, 15))

        # 🔥 TABLE DATA
        data = [[
            "Reg No", "Name", "Gen", "Age", "City",
            "Course", "Spec", "Yr", "Sem",
            "Att%", "Asg", "Int", "Prac",
            "Final", "Tot", "%", "Gr"
        ]]

        for s in students:
            data.append([
                s["registration_no"],
                s["student_name"][:12],  # 🔥 trim long names
                s["gender"],
                s["age"],
                s["city"][:10],
                s["course"][:10],
                s["specialization"][:10],
                s["year"],
                s["semester"],
                f'{s["attendance_percentage"]}%',
                s["assignment_marks"],
                s["internal_marks"],
                s["practical_marks"],
                s["final_exam_marks"],
                s["total_marks"],
                f'{s["percentage"]}%',
                s["grade"]
            ])

        # 🔥 FIXED WIDTH (VERY IMPORTANT)
        col_widths = [
            60, 80, 35, 30, 60,
            70, 70, 30, 30,
            45, 35, 35, 35,
            40, 45, 40, 35
        ]

        table = Table(data, repeatRows=1, colWidths=col_widths)

        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.darkblue),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),

            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),

            ("FONTSIZE", (0, 0), (-1, -1), 7),

            ("BOTTOMPADDING", (0, 0), (-1, 0), 6),

            ("BACKGROUND", (0, 1), (-1, -1), colors.beige),

            ("GRID", (0, 0), (-1, -1), 0.25, colors.grey)
        ]))

        elements.append(table)

        # 🔥 FOOTER
        def footer(canvas, doc):
            canvas.saveState()
            canvas.setFont("Helvetica", 8)

            canvas.drawString(
                30,
                15,
                "Generated by Student Performance Analytics System"
            )

            canvas.drawRightString(
                doc.pagesize[0] - 30,
                15,
                f"Page {doc.page}"
            )

            canvas.restoreState()

        doc.build(elements, onFirstPage=footer, onLaterPages=footer)

        return send_file(pdf_path, as_attachment=True)

    except Exception as e:
        return f"Error generating PDF: {str(e)}"

@app.route("/export_excel")
def export_excel():

    # 🔐 Auth check
    if "username" not in session:
        return redirect("/login")

    if session.get("role") != "admin":
        return redirect("/student_dashboard")

    try:
        conn = get_connection()
        cursor = conn.cursor(dictionary=True)

        cursor.execute("""
            SELECT
                registration_no, student_name, gender, age, city,
                course, specialization, year, semester,
                attendance_percentage, assignment_marks,
                internal_marks, practical_marks, final_exam_marks,
                total_marks, percentage, grade
            FROM students
            ORDER BY student_id
        """)

        students = cursor.fetchall()

        cursor.close()
        conn.close()

        # 🔥 Unique filename
        filename = f"student_report_{int(datetime.now().timestamp())}.xlsx"
        excel_path = os.path.join("static", filename)

        wb = Workbook()
        ws = wb.active
        ws.title = "Student Report"

        # ---------------- Title ----------------
        ws["A1"] = "Student Performance Analytics System"
        ws["A2"] = "Complete Student Report"

        ws["A1"].font = Font(size=18, bold=True)
        ws["A2"].font = Font(size=14, bold=True)

        ws["A4"] = "Generated On :"
        ws["B4"] = datetime.now().strftime("%d-%m-%Y %I:%M %p")

        ws["A5"] = "Total Students :"
        ws["B5"] = len(students)

        # ---------------- Header ----------------
        headers = [
            "Registration No", "Student Name", "Gender", "Age", "City",
            "Course", "Specialization", "Year", "Semester",
            "Attendance %", "Assignment", "Internal", "Practical",
            "Final Exam", "Total Marks", "Percentage", "Grade"
        ]

        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        start_row = 7

        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=start_row, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

        # 🔥 Freeze header row
        ws.freeze_panes = "A8"

        # ---------------- Data ----------------
        row = start_row + 1

        for s in students:
            values = [
                s["registration_no"], s["student_name"], s["gender"], s["age"], s["city"],
                s["course"], s["specialization"], s["year"], s["semester"],
                s["attendance_percentage"], s["assignment_marks"], s["internal_marks"],
                s["practical_marks"], s["final_exam_marks"], s["total_marks"],
                s["percentage"], s["grade"]
            ]

            for col, value in enumerate(values, start=1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border

            row += 1

        # ---------------- Auto Column Width ----------------
        for column in ws.columns:
            max_length = 0
            col_letter = get_column_letter(column[0].column)

            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass

            ws.column_dimensions[col_letter].width = max_length + 3

        # ---------------- Save ----------------
        wb.save(excel_path)

        return send_file(excel_path, as_attachment=True)

    except Exception as e:
        return f"Error generating Excel: {str(e)}"
    
@app.route("/student/<int:student_id>")
def student_details(student_id):

    # 🔐 Auth check
    if "username" not in session:
        return redirect("/login")

    try:
        conn = get_connection()
        cursor = conn.cursor(dictionary=True)

        cursor.execute(
            "SELECT * FROM students WHERE student_id = %s",
            (student_id,)
        )

        student = cursor.fetchone()

        # ❌ If student not found
        if not student:
            flash("Student not found.", "danger")
            return redirect("/students")

        # 🔐 OPTIONAL (Recommended)
        # Agar student role hai toh sirf apna data dekhe
        if session.get("role") == "student":
            if student["student_id"] != session.get("student_id"):
                flash("Unauthorized access.", "danger")
                return redirect("/student_dashboard")

        # 🔥 Reusable function
        def get_status(value, rules):
            for condition, label, color in rules:
                if condition(value):
                    return label, color

        # 🎯 Performance Rules
        performance_status, performance_color = get_status(
            student["percentage"],
            [
                (lambda x: x >= 90, "Excellent Performer", "#16a34a"),
                (lambda x: x >= 75, "Very Good Performer", "#2563eb"),
                (lambda x: x >= 60, "Good Performer", "#ca8a04"),
                (lambda x: x >= 40, "Average Performer", "#ea580c"),
                (lambda x: True, "Needs Improvement", "#dc2626"),
            ]
        )

        # 📊 Attendance Rules
        attendance_status, attendance_color = get_status(
            student["attendance_percentage"],
            [
                (lambda x: x >= 90, "Excellent Attendance", "#16a34a"),
                (lambda x: x >= 75, "Good Attendance", "#2563eb"),
                (lambda x: x >= 60, "Average Attendance", "#ca8a04"),
                (lambda x: True, "Low Attendance", "#dc2626"),
            ]
        )

        return render_template(
            "student_details.html",
            student=student,
            performance_status=performance_status,
            performance_color=performance_color,
            attendance_status=attendance_status,
            attendance_color=attendance_color
        )

    except Exception as e:
        return f"Error: {str(e)}"

    finally:
        cursor.close()
        conn.close()
    
@app.route("/add_student", methods=["GET", "POST"])
def add_student():

    # 🔐 Auth check
    if "username" not in session:
        return redirect("/login")

    if session.get("role") != "admin":
        return redirect("/student_dashboard")

    if request.method == "POST":

        # 🔹 Get form data
        registration_no = request.form.get("registration_no", "").strip()
        student_name = request.form.get("student_name", "").strip()
        gender = request.form.get("gender", "")
        age = request.form.get("age", "")
        city = request.form.get("city", "").strip()
        course = request.form.get("course", "").strip()
        specialization = request.form.get("specialization", "").strip()
        year = request.form.get("year", "")
        semester = request.form.get("semester", "")

        # 🔐 Basic validation
        if not all([registration_no, student_name, gender, age, course]):
            flash("Please fill all required fields.", "danger")
            return redirect("/add_student")

        try:
            conn = get_connection()
            cursor = conn.cursor(dictionary=True)

            # 🔍 Check duplicate
            cursor.execute(
                "SELECT student_id FROM students WHERE registration_no = %s",
                (registration_no,)
            )
            if cursor.fetchone():
                flash("Registration number already exists.", "danger")
                return redirect("/add_student")

            # 📝 Insert student
            cursor.execute("""
                INSERT INTO students (
                    registration_no, student_name, gender, age, city,
                    course, specialization, year, semester,
                    attendance_percentage, assignment_marks,
                    internal_marks, practical_marks, final_exam_marks,
                    total_marks, percentage, grade
                )
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,0,0,0,0,0,0,0,'N/A')
            """, (
                registration_no, student_name, gender, age, city,
                course, specialization, year, semester
            ))

            student_id = cursor.lastrowid

            # 🔐 Create login
            hashed_password = generate_password_hash("student123")

            cursor.execute("""
                INSERT INTO users (username, password, role, student_id)
                VALUES (%s, %s, 'student', %s)
            """, (
                registration_no,
                hashed_password,
                student_id
            ))

            conn.commit()

            log_activity(
                session["username"],
                "Add Student",
                f"Added Student: {registration_no} - {student_name}"
            )

            flash("Student added successfully.", "success")
            return redirect("/students")

        except Exception as e:
            conn.rollback()
            return f"Error: {str(e)}"

        finally:
            cursor.close()
            conn.close()

    # 🔢 Generate new registration number
    conn = get_connection()
    cursor = conn.cursor(dictionary=True)

    cursor.execute("""
        SELECT registration_no
        FROM students
        ORDER BY student_id DESC
        LIMIT 1
    """)

    last_student = cursor.fetchone()

    if last_student:
        last_number = int(last_student["registration_no"].replace("SPAI", ""))
        new_registration_no = f"SPAI{last_number + 1}"
    else:
        new_registration_no = "SPAI20260001"

    cursor.close()
    conn.close()

    return render_template(
        "add_student.html",
        new_registration_no=new_registration_no
    )
    
@app.route("/edit_student/<int:student_id>", methods=["GET", "POST"])
def edit_student(student_id):

    # 🔐 Auth check
    if "username" not in session:
        return redirect("/login")

    if session.get("role") != "admin":
        return redirect("/student_dashboard")

    try:
        page = int(request.args.get("page", 1))
    except:
        page = 1

    search = request.args.get("search", "")
    course_filter = request.args.get("course", "")   # ⚠️ name clash avoid

    conn = None
    cursor = None

    try:
        conn = get_connection()
        cursor = conn.cursor(dictionary=True)

        # 🔍 Fetch student
        cursor.execute(
            "SELECT * FROM students WHERE student_id = %s",
            (student_id,)
        )
        student = cursor.fetchone()

        if not student:
            flash("Student not found.", "danger")
            return redirect(url_for("students", page=page, search=search, course=course_filter))

        # ================= POST =================
        if request.method == "POST":

            registration_no = request.form.get("registration_no", "").strip()
            student_name = request.form.get("student_name", "").strip()
            gender = request.form.get("gender", "")
            age = request.form.get("age", "")
            city = request.form.get("city", "").strip()
            course = request.form.get("course", "").strip()
            specialization = request.form.get("specialization", "").strip()
            year = request.form.get("year", "")
            semester = request.form.get("semester", "")

            # 🔐 Validation
            if not all([registration_no, student_name, gender, age, course]):
                flash("Please fill all required fields.", "danger")
                return redirect(url_for(
                    "edit_student",
                    student_id=student_id,
                    page=page,
                    search=search,
                    course=course_filter
                ))

            # 🔍 Duplicate check (students)
            cursor.execute("""
                SELECT student_id FROM students
                WHERE registration_no = %s AND student_id != %s
            """, (registration_no, student_id))

            if cursor.fetchone():
                flash("Registration number already exists.", "danger")
                return redirect(url_for(
                    "edit_student",
                    student_id=student_id,
                    page=page,
                    search=search,
                    course=course_filter
                ))

            # 🔍 Duplicate check (users)
            cursor.execute("""
                SELECT * FROM users
                WHERE username = %s AND student_id != %s
            """, (registration_no, student_id))

            if cursor.fetchone():
                flash("Username already exists.", "danger")
                return redirect(url_for(
                    "edit_student",
                    student_id=student_id,
                    page=page,
                    search=search,
                    course=course_filter
                ))

            # 📝 Update students
            cursor.execute("""
                UPDATE students
                SET registration_no=%s,
                    student_name=%s,
                    gender=%s,
                    age=%s,
                    city=%s,
                    course=%s,
                    specialization=%s,
                    year=%s,
                    semester=%s
                WHERE student_id=%s
            """, (
                registration_no,
                student_name,
                gender,
                age,
                city,
                course,
                specialization,
                year,
                semester,
                student_id
            ))

            # 📝 Update users
            cursor.execute("""
                UPDATE users
                SET username = %s
                WHERE student_id = %s
            """, (registration_no, student_id))

            conn.commit()

            flash("Student updated successfully!", "success")

            # ✅ IMPORTANT: pagination maintain
            return redirect(url_for(
                "students",
                page=page,
                search=search,
                course=course_filter
            ))

        # ================= GET =================
        return render_template(
            "edit_student.html",
            student=student,
            page=page,
            search=search,
            course=course_filter
        )

    except Exception as e:
        if conn:
            conn.rollback()
        print("Edit Error:", e)
        flash("Error updating student.", "danger")
        return redirect(url_for("students", page=page, search=search, course=course_filter))

    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()
    
@app.route("/delete_student/<int:student_id>")
def delete_student(student_id):

    # 🔐 Auth check
    if "username" not in session:
        return redirect("/login")

    if session.get("role") != "admin":
        return redirect("/student_dashboard")

    try:
        page = int(request.args.get("page", 1))
    except:
        page = 1

    search = request.args.get("search", "")
    course = request.args.get("course", "")

    conn = None
    cursor = None

    try:
        conn = get_connection()
        cursor = conn.cursor()

        # 🔍 Fetch student for logging
        cursor.execute("""
            SELECT registration_no, student_name
            FROM students
            WHERE student_id = %s
        """, (student_id,))
        student = cursor.fetchone()

        # 🗑 Delete from users
        cursor.execute("""
            DELETE FROM users
            WHERE student_id = %s
        """, (student_id,))

        # 🗑 Delete from students
        cursor.execute("""
            DELETE FROM students
            WHERE student_id = %s
        """, (student_id,))

        conn.commit()

        # 🧾 Logging
        if student:
            log_activity(
                session["username"],
                "Delete Student",
                f"{student[0]} - {student[1]}"
            )

        flash("Student deleted successfully!", "success")

    except Exception as e:
        if conn:
            conn.rollback()
        print("Delete Error:", e)
        flash("Error deleting student.", "danger")

    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()
    return redirect(url_for(
        "students",
        page=page,
        search=search,
        course=course
    ))

@app.route("/delete_selected_students", methods=["POST"])
def delete_selected_students():

    if "username" not in session:
        return {"success": False}

    if session["role"] != "admin":
        return {"success": False}

    data = request.get_json()
    student_ids = data.get("student_ids", [])

    if not student_ids:
        return {
            "success": False,
            "message": "No students selected."
        }

    conn = None
    cursor = None

    try:
        conn = get_connection()
        cursor = conn.cursor()

        placeholders = ",".join(["%s"] * len(student_ids))

        # Fetch student details for logging
        cursor.execute(f"""
            SELECT registration_no, student_name
            FROM students
            WHERE student_id IN ({placeholders})
        """, tuple(student_ids))

        students = cursor.fetchall()

        # 1. Delete from users first
        cursor.execute(f"""
            DELETE FROM users
            WHERE student_id IN ({placeholders})
        """, tuple(student_ids))

        # 2. Delete from students
        cursor.execute(f"""
            DELETE FROM students
            WHERE student_id IN ({placeholders})
        """, tuple(student_ids))

        conn.commit()

        # Logging
        for student in students:
            log_activity(
                session["username"],
                "Bulk Delete",
                f"Registration No: {student[0]} | Name: {student[1]}"
            )

        return {
            "success": True,
            "deleted": len(students)
        }

    except Exception as e:
        if conn:
            conn.rollback()
        print("Bulk Delete Error:", e)

        return {
            "success": False,
            "message": "Error deleting students."
        }

    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()

@app.route("/student_dashboard")
def student_dashboard():

    # 🔐 Auth check
    if "username" not in session:
        return redirect("/login")

    if session.get("role") != "student":
        return redirect("/dashboard")

    try:
        conn = get_connection()
        cursor = conn.cursor(dictionary=True)

        cursor.execute("""
            SELECT s.*
            FROM students s
            JOIN users u
                ON s.student_id = u.student_id
            WHERE u.username = %s
        """, (session["username"],))

        student = cursor.fetchone()

        # ❌ Student not found
        if not student:
            flash("Student record not found.", "danger")
            return redirect("/login")

        # 📊 Progress bars (safe calculation)
        student["assignment_width"] = (student["assignment_marks"] or 0) * 10
        student["internal_width"] = ((student["internal_marks"] or 0) / 15) * 100
        student["practical_width"] = ((student["practical_marks"] or 0) / 15) * 100
        student["final_width"] = ((student["final_exam_marks"] or 0) / 60) * 100

        # 🔥 Reusable status function
        def get_status(value, rules):
            for condition, label, color in rules:
                if condition(value):
                    return label, color

        # 🎯 Performance
        performance_status, performance_color = get_status(
            student["percentage"],
            [
                (lambda x: x >= 90, "Excellent Performer", "#16a34a"),
                (lambda x: x >= 75, "Very Good Performer", "#2563eb"),
                (lambda x: x >= 60, "Good Performer", "#ca8a04"),
                (lambda x: x >= 40, "Average Performer", "#ea580c"),
                (lambda x: True, "Needs Improvement", "#dc2626"),
            ]
        )

        # 📅 Attendance
        attendance_status, attendance_color = get_status(
            student["attendance_percentage"],
            [
                (lambda x: x >= 90, "Excellent Attendance", "#16a34a"),
                (lambda x: x >= 75, "Good Attendance", "#2563eb"),
                (lambda x: x >= 60, "Average Attendance", "#ca8a04"),
                (lambda x: True, "Low Attendance", "#dc2626"),
            ]
        )

        return render_template(
            "student_dashboard.html",
            student=student,
            performance_status=performance_status,
            performance_color=performance_color,
            attendance_status=attendance_status,
            attendance_color=attendance_color
        )

    except Exception as e:
        return f"Error: {str(e)}"

    finally:
        cursor.close()
        conn.close()

@app.route("/change_password", methods=["GET", "POST"])
def change_password():

    # 🔐 Auth check
    if "username" not in session:
        return redirect("/login")

    if request.method == "POST":

        old_password = request.form.get("old_password", "")
        new_password = request.form.get("new_password", "")
        confirm_password = request.form.get("confirm_password", "")

        # 🔐 Validation
        if not all([old_password, new_password, confirm_password]):
            flash("All fields are required.", "danger")
            return redirect("/change_password")

        if new_password != confirm_password:
            flash("New Password and Confirm Password do not match.", "danger")
            return redirect("/change_password")

        if len(new_password) < 8:
            flash("Password must be at least 8 characters.", "danger")
            return redirect("/change_password")

        if not re.search(r"[A-Z]", new_password):
            flash("Password must contain at least one uppercase letter.", "danger")
            return redirect("/change_password")

        if not re.search(r"[a-z]", new_password):
            flash("Password must contain at least one lowercase letter.", "danger")
            return redirect("/change_password")

        if not re.search(r"\d", new_password):
            flash("Password must contain at least one number.", "danger")
            return redirect("/change_password")

        if not re.search(r"[!@#$%^&*(),.?\":{}|<>]", new_password):
            flash("Password must contain at least one special character.", "danger")
            return redirect("/change_password")

        try:
            conn = get_connection()
            cursor = conn.cursor(dictionary=True)

            # 🔍 Fetch user
            cursor.execute(
                "SELECT * FROM users WHERE username = %s",
                (session["username"],)
            )
            user = cursor.fetchone()

            if not user:
                flash("User not found.", "danger")
                return redirect("/login")

            # 🔐 Check old password
            if not check_password_hash(user["password"], old_password):
                flash("Old Password is incorrect!", "danger")
                return redirect("/change_password")

            # 🔐 Hash new password
            hashed_password = generate_password_hash(new_password)

            # 📝 Update password
            cursor.execute("""
                UPDATE users
                SET password = %s
                WHERE username = %s
            """, (hashed_password, session["username"]))

            conn.commit()

            # 🧾 Log activity
            log_activity(
                session["username"],
                "Change Password",
                "User changed password successfully"
            )

            # 🔥 Security: logout after password change
            session.clear()

            flash("Password changed successfully. Please login again.", "success")
            return redirect("/login")

        except Exception as e:
            conn.rollback()
            return f"Error: {str(e)}"

        finally:
            cursor.close()
            conn.close()

    return render_template("change_password.html")

from werkzeug.security import generate_password_hash

@app.route("/reset_password/<int:student_id>", methods=["POST"])
def reset_password(student_id):

    # 🔐 Auth check
    if "username" not in session:
        return redirect("/login")

    if session.get("role") != "admin":
        return redirect("/student_dashboard")

    conn = None
    cursor = None

    try:
        conn = get_connection()
        cursor = conn.cursor(dictionary=True)

        # 🔍 Fetch student
        cursor.execute("""
            SELECT registration_no, student_name
            FROM students
            WHERE student_id = %s
        """, (student_id,))

        student = cursor.fetchone()

        if not student:
            flash("Student not found.", "danger")
            return redirect("/students")

        # 🔐 FIXED PASSWORD
        default_password = "student123"
        hashed_password = generate_password_hash(default_password)

        # 📝 Update password
        cursor.execute("""
            UPDATE users
            SET password = %s
            WHERE student_id = %s
        """, (hashed_password, student_id))

        conn.commit()

        # 🧾 Log
        log_activity(
            session["username"],
            "Reset Password",
            f"Reset password for: {student['registration_no']} - {student['student_name']}"
        )

        flash(
            "Password reset successful. Default Password: student123",
            "success"
        )

        return redirect("/students")

    except Exception as e:
        if conn:
            conn.rollback()
        return f"Error: {str(e)}"

    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()

def get_logged_in_student(username):
    conn = get_connection()
    cursor = conn.cursor(dictionary=True)

    cursor.execute("""
        SELECT s.*
        FROM students s
        JOIN users u
            ON s.student_id = u.student_id
        WHERE u.username = %s
    """, (username,))

    student = cursor.fetchone()

    cursor.close()
    conn.close()

    return student

@app.route("/my_profile")
def my_profile():

    if "username" not in session:
        return redirect("/login")

    if session.get("role") != "student":
        return redirect("/dashboard")

    student = get_logged_in_student(session["username"])

    if not student:
        flash("Student data not found.", "danger")
        return redirect("/login")

    return render_template(
        "my_profile.html",
        student=student
    )
    
@app.route("/my_performance")
def my_performance():

    if "username" not in session:
        return redirect("/login")

    if session.get("role") != "student":
        return redirect("/dashboard")

    student = get_logged_in_student(session["username"])

    if not student:
        flash("Student data not found.", "danger")
        return redirect("/login")

    return render_template(
        "my_performance.html",
        student=student
    )
    
@app.route("/update_marks/<int:student_id>", methods=["GET", "POST"])
def update_marks(student_id):

    # 🔐 Auth check
    if "username" not in session:
        return redirect("/login")

    if session.get("role") != "admin":
        return redirect("/student_dashboard")

    try:
        page = int(request.args.get("page", 1))
    except:
        page = 1

    search = request.args.get("search", "")
    course = request.args.get("course", "")

    conn = None
    cursor = None

    try:
        conn = get_connection()
        cursor = conn.cursor(dictionary=True)

        # 🔍 Fetch student
        cursor.execute(
            "SELECT * FROM students WHERE student_id = %s",
            (student_id,)
        )
        student = cursor.fetchone()

        if not student:
            flash("Student not found.", "danger")
            return redirect(url_for("students"))

        # ================= POST =================
        if request.method == "POST":

            attendance = float(request.form.get("attendance_percentage", 0))
            assignment = float(request.form.get("assignment_marks", 0))
            internal = float(request.form.get("internal_marks", 0))
            practical = float(request.form.get("practical_marks", 0))
            final_exam = float(request.form.get("final_exam_marks", 0))

            total = assignment + internal + practical + final_exam
            percentage = total

            # 🎓 Grade logic
            if percentage >= 90:
                grade = "O"
            elif percentage >= 80:
                grade = "A+"
            elif percentage >= 70:
                grade = "A"
            elif percentage >= 60:
                grade = "B+"
            elif percentage >= 50:
                grade = "B"
            elif percentage >= 40:
                grade = "C"
            else:
                grade = "F"

            cursor.execute("""
                UPDATE students
                SET attendance_percentage=%s,
                    assignment_marks=%s,
                    internal_marks=%s,
                    practical_marks=%s,
                    final_exam_marks=%s,
                    total_marks=%s,
                    percentage=%s,
                    grade=%s
                WHERE student_id=%s
            """, (
                attendance,
                assignment,
                internal,
                practical,
                final_exam,
                total,
                percentage,
                grade,
                student_id
            ))

            conn.commit()

            flash("Marks updated successfully!", "success")
            return redirect(url_for("students", page=page, search=search, course=course))

        # ================= GET =================
        return render_template(
            "update_marks.html",
            student=student,
            page=page,
            search=search,
            course=course
        )

    except Exception as e:
        if conn:
            conn.rollback()
        return f"Error: {str(e)}"

    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()
    
@app.route("/import_students")
def import_students():

    # 🔐 Auth check
    if "username" not in session:
        return redirect("/login")

    if session.get("role") != "admin":
        return redirect("/student_dashboard")

    report = session.get("import_report", {})

    return render_template("import_students.html", report=report)
    
@app.route("/clear_import_report", methods=["GET", "POST"])
def clear_import_report():

    if "username" not in session:
        return redirect("/login")

    if session.get("role") != "admin":
        return redirect("/student_dashboard")

    session.pop("import_report", None)

    flash("Import report cleared.", "success")

    return redirect("/import_students")

@app.route("/upload_students", methods=["POST"])
def upload_students():

    if "username" not in session:
        return redirect("/login")

    if session.get("role") != "admin":
        return redirect("/student_dashboard")

    if "excel_file" not in request.files:
        flash("No file selected.", "danger")
        return redirect("/import_students")

    file = request.files["excel_file"]

    if file.filename == "":
        flash("Please choose an Excel file.", "danger")
        return redirect("/import_students")

    if not file.filename.lower().endswith(".xlsx"):
        flash("Only .xlsx files are allowed.", "danger")
        return redirect("/import_students")

    # 🔥 Unique filename
    filename = f"{uuid.uuid4().hex}.xlsx"
    filepath = os.path.join(app.config["UPLOAD_FOLDER"], filename)
    file.save(filepath)

    workbook = load_workbook(filepath)
    sheet = workbook["Students"]

    rows = list(sheet.iter_rows(values_only=True))

    expected_headers = [
        "Registration No","Student Name","Gender","Age","City",
        "Course","Specialization","Year","Semester",
        "Attendance","Assignment","Internal","Practical","Final"
    ]

    headers = [str(h).strip() if h else "" for h in rows[0]]

    if headers != expected_headers:
        flash("Invalid Excel Template!", "danger")
        return redirect("/import_students")

    conn = get_connection()
    cursor = conn.cursor()

    imported = 0
    duplicates = 0
    invalid = 0
    invalid_rows = []

    try:
        # 🔥 Fetch existing users + students
        cursor.execute("SELECT registration_no FROM students")
        existing_students = {row[0] for row in cursor.fetchall()}

        cursor.execute("SELECT username FROM users")
        existing_users = {row[0] for row in cursor.fetchall()}

        for excel_row, row in enumerate(rows[1:], start=2):
            try:
                if not row[0]:
                    continue

                reg = str(row[0]).strip()
                name = str(row[1]).strip()
                gender = str(row[2]).strip().capitalize()
                age = int(row[3])
                city = str(row[4]).strip()
                course = str(row[5]).strip()
                specialization = str(row[6]).strip()
                year = int(row[7])
                semester = int(row[8])

                attendance = float(row[9])
                assignment = float(row[10])
                internal = float(row[11])
                practical = float(row[12])
                final_exam = float(row[13])

                # 🔐 Validation
                if gender not in ["Male", "Female"]:
                    invalid += 1
                    invalid_rows.append(f"Row {excel_row}: Invalid gender")
                    continue

                if not (0 <= attendance <= 100 and
                        0 <= assignment <= 10 and
                        0 <= internal <= 15 and
                        0 <= practical <= 15 and
                        0 <= final_exam <= 60 and
                        1 <= year <= 4 and
                        1 <= semester <= 8):
                    
                    invalid += 1
                    invalid_rows.append(f"Row {excel_row}: Invalid marks")
                    continue

                # 🔍 Duplicate check
                if reg in existing_students or reg in existing_users:
                    duplicates += 1
                    continue

                # 🧮 Correct calculation
                total = assignment + internal + practical + final_exam
                percentage = total

                def grade_fn(p):
                    return (
                        "O" if p >= 90 else
                        "A+" if p >= 80 else
                        "A" if p >= 70 else
                        "B+" if p >= 60 else
                        "B" if p >= 50 else
                        "C" if p >= 40 else "F"
                    )

                grade = grade_fn(percentage)

                # 📝 Insert student
                cursor.execute("""
                    INSERT INTO students(
                        registration_no, student_name, gender, age, city,
                        course, specialization, year, semester,
                        attendance_percentage, assignment_marks,
                        internal_marks, practical_marks, final_exam_marks,
                        total_marks, percentage, grade
                    )
                    VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                """, (
                    reg, name, gender, age, city,
                    course, specialization, year, semester,
                    attendance, assignment, internal, practical,
                    final_exam, total, percentage, grade
                ))

                student_id = cursor.lastrowid  # 🔥 important

                # 🔐 Create login
                hashed_password = generate_password_hash("student123")

                cursor.execute("""
                    INSERT INTO users (username, password, role, student_id)
                    VALUES (%s, %s, 'student', %s)
                """, (
                    reg,
                    hashed_password,
                    student_id
                ))

                existing_students.add(reg)
                existing_users.add(reg)

                imported += 1

            except Exception:
                invalid += 1
                invalid_rows.append(f"Row {excel_row}: Error")
                continue

        conn.commit()

        log_activity(
            session["username"],
            "Import Excel",
            f"Imported: {imported}, Duplicate: {duplicates}, Invalid: {invalid}"
        )

    except Exception as e:
        conn.rollback()
        return f"Error: {str(e)}"

    finally:
        cursor.close()
        conn.close()
        if os.path.exists(filepath):
            os.remove(filepath)

    session["import_report"] = {
        "imported": imported,
        "duplicates": duplicates,
        "invalid": invalid,
        "errors": invalid_rows
    }

    return redirect("/import_students")

@app.route("/download_template")
def download_template():

    if "username" not in session:
        return redirect("/login")

    if session.get("role") != "admin":
        return redirect("/student_dashboard")

    wb = Workbook()

    # -----------------------------
    # Sheet 1: Students
    # -----------------------------
    ws = wb.active
    ws.title = "Students"

    headers = [
        "Registration No","Student Name","Gender","Age","City",
        "Course","Specialization","Year","Semester",
        "Attendance","Assignment","Internal","Practical","Final"
    ]

    header_fill = PatternFill(start_color="2F63E9", end_color="2F63E9", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Sample Row
    sample = [
        "REG001","Rahul Sharma","Male",20,"Delhi",
        "BCA","Data Science",2,4,92,8,13,14,54
    ]

    for col, value in enumerate(sample, start=1):
        ws.cell(row=2, column=col).value = value

    # -----------------------------
    # Data Validations
    # -----------------------------

    # Gender dropdown
    gender_validation = DataValidation(
        type="list",
        formula1='"Male,Female"',
        allow_blank=False
    )
    ws.add_data_validation(gender_validation)
    gender_validation.add("C2:C500")

    # Numeric validations
    validations = [
        ("J2:J500", 0, 100),  # Attendance
        ("K2:K500", 0, 10),   # Assignment
        ("L2:L500", 0, 15),   # Internal
        ("M2:M500", 0, 15),   # Practical
        ("N2:N500", 0, 60),   # Final
    ]

    for cell_range, min_val, max_val in validations:
        dv = DataValidation(
            type="decimal",
            operator="between",
            formula1=str(min_val),
            formula2=str(max_val),
            allow_blank=False
        )
        ws.add_data_validation(dv)
        dv.add(cell_range)

    # -----------------------------
    # Column Width
    # -----------------------------
    widths = {
        "A":18,"B":25,"C":15,"D":10,"E":20,
        "F":18,"G":20,"H":10,"I":12,
        "J":15,"K":15,"L":15,"M":15,"N":15
    }

    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    ws.freeze_panes = "A2"

    # 🔒 Protect header
    ws.protection.sheet = True

    # -----------------------------
    # Sheet 2: Instructions
    # -----------------------------
    instructions = wb.create_sheet("Instructions")

    instructions["A1"] = "Student Import Instructions"
    instructions["A1"].font = Font(bold=True, size=16, color="FFFFFF")
    instructions["A1"].fill = PatternFill(start_color="2F63E9", end_color="2F63E9", fill_type="solid")

    rules = [
        "Do NOT rename any column headings.",
        "Registration Number must be unique.",
        "Gender must be Male or Female.",
        "Attendance must be between 0 and 100.",
        "Assignment Marks must be between 0 and 10.",
        "Internal Marks must be between 0 and 15.",
        "Practical Marks must be between 0 and 15.",
        "Final Exam Marks must be between 0 and 60.",
        "Year must be between 1 and 4.",
        "Semester must be between 1 and 8.",
        "Do not leave mandatory fields blank.",
        "Upload only .xlsx files."
    ]

    for i, rule in enumerate(rules, start=3):
        instructions[f"A{i}"] = f"{i-2}. {rule}"
        instructions[f"A{i}"].alignment = Alignment(wrap_text=True)

    instructions.column_dimensions["A"].width = 100

    # -----------------------------
    # Save to Memory
    # -----------------------------
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        download_name="Student_Import_Template.xlsx",
        as_attachment=True,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    
@app.route("/analytics")
def analytics():

    if "username" not in session:
        return redirect("/login")

    return render_template("analytics.html")

if __name__ == "__main__":
    app.run(debug=True)